
import customtkinter as ctk
import sqlite3
import csv
from tkinter import messagebox, filedialog

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

class OgrenciYonetim:
    def __init__(self, root):
        self.root = root
        self.root.title("Öğrenci Yönetim Sistemi")
        self.root.geometry("600x650")
        self.db_baglanti()
        self.title_label = ctk.CTkLabel(root, text="Öğrenci Yönetim Sistemi", font=ctk.CTkFont(size=22, weight="bold"))
        self.title_label.pack(pady=10)
        self.main_frame = ctk.CTkFrame(root)
        self.main_frame.pack(pady=10, padx=20, fill="both", expand=True)
        self.form_frame = ctk.CTkFrame(self.main_frame)
        self.form_frame.pack(pady=10, padx=10, fill="x")
        self.label_numara = ctk.CTkLabel(self.form_frame, text="Numara:")
        self.label_numara.grid(row=0, column=0, padx=10, pady=5, sticky="e")
        self.entry_numara = ctk.CTkEntry(self.form_frame, width=250)
        self.entry_numara.grid(row=0, column=1, padx=10, pady=5)
        self.label_ad = ctk.CTkLabel(self.form_frame, text="Ad:")
        self.label_ad.grid(row=1, column=0, padx=10, pady=5, sticky="e")
        self.entry_ad = ctk.CTkEntry(self.form_frame, width=250)
        self.entry_ad.grid(row=1, column=1, padx=10, pady=5)
        self.label_soyad = ctk.CTkLabel(self.form_frame, text="Soyad:")
        self.label_soyad.grid(row=2, column=0, padx=10, pady=5, sticky="e")
        self.entry_soyad = ctk.CTkEntry(self.form_frame, width=250)
        self.entry_soyad.grid(row=2, column=1, padx=10, pady=5)
        self.btn_ekle = ctk.CTkButton(self.form_frame, text="Ekle / Güncelle", command=self.ogrenci_ekle_guncelle)
        self.btn_ekle.grid(row=3, column=0, columnspan=2, pady=10)
        self.search_frame = ctk.CTkFrame(self.main_frame)
        self.search_frame.pack(pady=5, padx=10, fill="x")
        self.search_entry = ctk.CTkEntry(self.search_frame, placeholder_text="Ara (Ad, Soyad veya Numara)", width=300)
        self.search_entry.pack(side="left", padx=(10, 5), pady=5)
        self.search_btn = ctk.CTkButton(self.search_frame, text="Ara", command=self.ogrenci_ara)
        self.search_btn.pack(side="left", padx=(5, 5), pady=5)
        self.clear_btn = ctk.CTkButton(self.search_frame, text="Temizle", command=self.ogrenci_listesi)
        self.clear_btn.pack(side="left", padx=(5, 10), pady=5)
        self.list_frame = ctk.CTkFrame(self.main_frame)
        self.list_frame.pack(pady=10, padx=10, fill="x")
        self.label_sec = ctk.CTkLabel(self.list_frame, text="Kayıtlı Öğrenciler:")
        self.label_sec.grid(row=0, column=0, padx=10, pady=5, sticky="e")
        self.optionmenu = ctk.CTkOptionMenu(self.list_frame, values=["Henüz öğrenci yok"])
        self.optionmenu.grid(row=0, column=1, padx=10, pady=5, sticky="w")
        self.btn_sil = ctk.CTkButton(self.list_frame, text="Seçili Öğrenciyi Sil", command=self.ogrenci_sil)
        self.btn_sil.grid(row=1, column=0, columnspan=2, pady=10)
        self.btn_export = ctk.CTkButton(self.main_frame, text="CSV Aktar", command=self.disari_aktar)
        self.btn_export.pack(pady=5)
        self.btn_excel = ctk.CTkButton(self.main_frame, text="Excel'e Aktar", command=self.excel_aktar)
        self.btn_excel.pack(pady=5)
        self.btn_pdf = ctk.CTkButton(self.main_frame, text="PDF'ye Aktar", command=self.pdf_aktar)
        self.btn_pdf.pack(pady=5)
        self.btn_login = ctk.CTkButton(self.main_frame, text="Giriş Paneli", command=self.kullanici_giris)
        self.btn_login.pack(pady=5)
        self.ogrenci_listesi()
        self.root.protocol("WM_DELETE_WINDOW", self.kapat)

    def db_baglanti(self):
        self.conn = sqlite3.connect("ogrenciler.db")
        self.cursor = self.conn.cursor()
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS ogrenciler (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numara TEXT UNIQUE,
                ad TEXT,
                soyad TEXT
            )
        """)
        self.conn.commit()

    def ogrenci_ekle_guncelle(self):
        numara = self.entry_numara.get()
        ad = self.entry_ad.get()
        soyad = self.entry_soyad.get()
        if not numara or not ad or not soyad:
            messagebox.showerror("Hata", "Lütfen tüm alanları doldurun!")
            return
        self.cursor.execute("SELECT numara FROM ogrenciler WHERE numara = ?", (numara,))
        if self.cursor.fetchone():
            self.cursor.execute("UPDATE ogrenciler SET ad=?, soyad=? WHERE numara=?", (ad, soyad, numara))
            messagebox.showinfo("Güncellendi", "Öğrenci bilgileri güncellendi.")
        else:
            try:
                self.cursor.execute("INSERT INTO ogrenciler (numara, ad, soyad) VALUES (?, ?, ?)", (numara, ad, soyad))
                messagebox.showinfo("Eklendi", "Öğrenci başarıyla eklendi.")
            except sqlite3.IntegrityError:
                messagebox.showerror("Hata", "Bu numara zaten kayıtlı!")
        self.conn.commit()
        self.entry_numara.delete(0, "end")
        self.entry_ad.delete(0, "end")
        self.entry_soyad.delete(0, "end")
        self.ogrenci_listesi()

    def ogrenci_listesi(self):
        self.cursor.execute("SELECT numara, ad, soyad FROM ogrenciler")
        rows = self.cursor.fetchall()
        if rows:
            ogrenciler = [f"{row[0]} - {row[1]} {row[2]}" for row in rows]
        else:
            ogrenciler = ["Henüz öğrenci yok"]
        self.optionmenu.configure(values=ogrenciler)
        self.optionmenu.set(ogrenciler[0])

    def ogrenci_sil(self):
        secili = self.optionmenu.get()
        if secili == "Henüz öğrenci yok":
            messagebox.showwarning("Uyarı", "Silinecek öğrenci yok!")
            return
        secili_numara = secili.split(" - ")[0]
        self.cursor.execute("DELETE FROM ogrenciler WHERE numara = ?", (secili_numara,))
        self.conn.commit()
        messagebox.showinfo("Silindi", "Öğrenci silindi.")
        self.ogrenci_listesi()

    def ogrenci_ara(self):
        sorgu = self.search_entry.get().strip().lower()
        self.cursor.execute("SELECT numara, ad, soyad FROM ogrenciler")
        filtrelenmis = [
            f"{row[0]} - {row[1]} {row[2]}"
            for row in self.cursor.fetchall()
            if sorgu in row[0].lower() or sorgu in row[1].lower() or sorgu in row[2].lower()
        ]
        if not filtrelenmis:
            filtrelenmis = ["Sonuç bulunamadı"]
        self.optionmenu.configure(values=filtrelenmis)
        self.optionmenu.set(filtrelenmis[0])

    def disari_aktar(self):
        self.cursor.execute("SELECT * FROM ogrenciler")
        rows = self.cursor.fetchall()
        if not rows:
            messagebox.showinfo("Bilgi", "Dışa aktarılacak veri yok.")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV dosyası", "*.csv")])
        if file_path:
            with open(file_path, mode='w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "Numara", "Ad", "Soyad"])
                writer.writerows(rows)
            messagebox.showinfo("Başarılı", "Veriler CSV olarak kaydedildi.")

    def excel_aktar(self):
        try:
            from openpyxl import Workbook
        except ImportError:
            messagebox.showerror("Eksik Kütüphane", "Lütfen 'openpyxl' yükleyin: pip install openpyxl")
            return
        self.cursor.execute("SELECT * FROM ogrenciler")
        rows = self.cursor.fetchall()
        if not rows:
            messagebox.showinfo("Bilgi", "Veri bulunamadı.")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel dosyası", "*.xlsx")])
        if file_path:
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "Numara", "Ad", "Soyad"])
            for row in rows:
                ws.append(row)
            wb.save(file_path)
            messagebox.showinfo("Başarılı", "Excel dosyası oluşturuldu.")

    def pdf_aktar(self):
        try:
            from fpdf import FPDF
        except ImportError:
            messagebox.showerror("Eksik Kütüphane", "Lütfen 'fpdf' yükleyin: pip install fpdf")
            return
        self.cursor.execute("SELECT numara, ad, soyad FROM ogrenciler")
        rows = self.cursor.fetchall()
        if not rows:
            messagebox.showinfo("Bilgi", "Veri bulunamadı.")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF dosyası", "*.pdf")])
        if file_path:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, txt="Öğrenci Listesi", ln=True, align="C")
            pdf.ln(10)
            for row in rows:
                pdf.cell(200, 10, txt=f"{row[0]} - {row[1]} {row[2]}", ln=True)
            pdf.output(file_path)
            messagebox.showinfo("Başarılı", "PDF dosyası oluşturuldu.")

    def kullanici_giris(self):
        login_win = ctk.CTkToplevel(self.root)
        login_win.title("Giriş Yap")
        login_win.geometry("300x200")
        login_win.grab_set()
        ctk.CTkLabel(login_win, text="Kullanıcı Adı:").pack(pady=5)
        entry_user = ctk.CTkEntry(login_win)
        entry_user.pack(pady=5)
        ctk.CTkLabel(login_win, text="Şifre:").pack(pady=5)
        entry_pass = ctk.CTkEntry(login_win, show="*")
        entry_pass.pack(pady=5)

        def check_credentials():
            if entry_user.get() == "admin" and entry_pass.get() == "1234":
                messagebox.showinfo("Giriş Başarılı", "Hoş geldiniz!")
                login_win.destroy()
            else:
                messagebox.showerror("Hatalı Giriş", "Kullanıcı adı veya şifre yanlış.")

        ctk.CTkButton(login_win, text="Giriş", command=check_credentials).pack(pady=10)

    def kapat(self):
        self.conn.close()
        self.root.destroy()

if __name__ == "__main__":
    root = ctk.CTk()
    app = OgrenciYonetim(root)
    root.mainloop()
